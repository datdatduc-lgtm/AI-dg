#!/usr/bin/env python3
"""Create AI-dg project material-summary and quotation workbooks.

This exporter is intentionally project/workspace-oriented. It reads whatever
verified/partial AI-dg JSON artifacts exist under OUTPUT/TAKEOFF and WORK/
geometry, then ALWAYS writes review-friendly Excel deliverables.

It never invents missing prices, quantities, suppliers, materials or images.
Missing values are left blank and surfaced with explicit status/review text.

Expected outputs:
  OUTPUT/EXCEL/AI-dg_Tong-hop-vat-lieu.xlsx
  OUTPUT/EXCEL/AI-dg_Bao-gia.xlsx

Optional image inputs:
  OUTPUT/IMAGES/<item>_iso.png
  OUTPUT/IMAGES/<item>_front.png
  OUTPUT/IMAGES/<item>_side.png

Usage:
  python export_project_excel.py D:/AI-dg/Project-A
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - explicit runtime dependency
    raise SystemExit(
        "Excel exporter requires openpyxl. Install AI-dg runtime extras or run: "
        "python -m pip install openpyxl Pillow"
    ) from exc

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
BOLD_FONT = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export AI-dg project Excel deliverables")
    parser.add_argument("project_root", type=Path)
    return parser.parse_args()


def load_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return {"_load_error": f"{path}: {exc}", "_default": default}


def payload_rows(payload: Any, keys: Iterable[str]) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if not isinstance(payload, dict):
        return []
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return [row for row in value if isinstance(row, dict)]
    return []


def slug(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def first_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, "", []):
            return value
    return None


def source_text(source: Any) -> str:
    if isinstance(source, str):
        return source
    if not isinstance(source, dict):
        return ""
    parts = []
    for key in ("pdf", "file", "relative_path", "page", "view", "evidence"):
        value = source.get(key)
        if value not in (None, ""):
            parts.append(f"{key}={value}")
    return " | ".join(parts)


def text_join(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(text_join(v) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={text_join(v)}" for k, v in value.items())
    return str(value)


def write_table(ws, headers: list[str], rows: list[list[Any]], start_row: int = 1) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    for r_index, row in enumerate(rows, start=start_row + 1):
        for c_index, value in enumerate(row, start=1):
            cell = ws.cell(r_index, c_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TABLE_BORDER
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    ws.freeze_panes = f"A{start_row + 1}"


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def image_index(images_root: Path) -> list[Path]:
    if not images_root.is_dir():
        return []
    return sorted(
        p for p in images_root.rglob("*")
        if p.is_file() and p.suffix.lower() in {".png", ".jpg", ".jpeg"}
    )


def find_item_image(images: list[Path], identifiers: list[Any]) -> Path | None:
    tokens = [slug(v) for v in identifiers if slug(v)]
    if not tokens:
        return None
    candidates = []
    for path in images:
        stem = slug(path.stem)
        if any(token in stem for token in tokens):
            priority = 0 if "ISO" in path.stem.upper() else 1
            candidates.append((priority, len(path.name), path))
    if not candidates:
        return None
    return sorted(candidates, key=lambda x: (x[0], x[1], str(x[2])))[0][2]


def add_image(ws, path: Path, anchor: str, max_width: int = 220, max_height: int = 120) -> bool:
    try:
        img = XLImage(str(path))
    except Exception:
        return False
    width = float(img.width or max_width)
    height = float(img.height or max_height)
    scale = min(max_width / width, max_height / height, 1.0)
    img.width = int(width * scale)
    img.height = int(height * scale)
    ws.add_image(img, anchor)
    return True


def unique_item_rows(items: list[dict[str, Any]], geometry_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in [*geometry_rows, *items]:
        item_id = first_value(row, "item_code", "item_id", "id", "code", "name")
        if not item_id:
            continue
        key = str(item_id)
        merged = result.setdefault(key, {})
        for field, value in row.items():
            if value not in (None, "", []):
                merged[field] = value
        merged.setdefault("_item_key", key)
    return list(result.values())


def material_key(row: dict[str, Any]) -> str:
    return str(first_value(row, "material_code", "source_material_code", "code", "material_name", "material") or "UNSPECIFIED")


def build_material_workbook(
    project_root: Path,
    project_meta: dict[str, Any],
    items: list[dict[str, Any]],
    geometry_rows: list[dict[str, Any]],
    material_regions: list[dict[str, Any]],
    bom_rows: list[dict[str, Any]],
    suppliers: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
    images: list[Path],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TONG_QUAN"
    ws["A1"] = "AI-dg — TỔNG HỢP VẬT LIỆU"
    ws["A1"].font = TITLE_FONT
    summary_rows = [
        ["Công trình", project_meta.get("project_name", project_root.name)],
        ["Project root", str(project_root)],
        ["Ngày xuất", datetime.now(timezone.utc).isoformat()],
        ["Số hạng mục", len(unique_item_rows(items, geometry_rows))],
        ["Số nhóm BOM", len(bom_rows)],
        ["Số nguồn cung cấp", len(suppliers)],
        ["Số review", len(review_rows)],
        ["Nguyên tắc", "Không tự bịa kích thước, vật liệu, số lượng, giá hoặc nhà cung cấp"],
    ]
    write_table(ws, ["Thông tin", "Giá trị"], summary_rows, start_row=3)
    set_widths(ws, {"A": 28, "B": 90})

    items_ws = wb.create_sheet("HANG_MUC")
    item_rows = unique_item_rows(items, geometry_rows)
    headers = ["Ảnh mô phỏng", "Mã", "Tên", "Kích thước / envelope", "Readiness", "Ruby", "Nguồn / ghi chú"]
    write_table(items_ws, headers, [], start_row=1)
    items_ws.freeze_panes = "A2"
    set_widths(items_ws, {"A": 32, "B": 16, "C": 28, "D": 35, "E": 24, "F": 42, "G": 70})

    ruby_root = project_root / "OUTPUT" / "RUBY"
    for index, row in enumerate(item_rows, start=2):
        item_code = first_value(row, "item_code", "item_id", "id", "code", "_item_key")
        item_name = first_value(row, "item_name", "name", "title")
        envelope = first_value(row, "envelope", "overall_envelope", "dimensions")
        readiness = first_value(row, "component_geometry_readiness", "readiness", "status")
        source = first_value(row, "source", "sources", "notes")
        image_path = find_item_image(images, [item_code, row.get("id"), item_name])

        ruby_path = None
        if ruby_root.is_dir():
            token = slug(item_code)
            matches = [p for p in ruby_root.glob("*.rb") if token and token in slug(p.stem)]
            if matches:
                ruby_path = matches[0]

        values = [
            "" if image_path else "CHƯA CÓ ẢNH — chạy Ruby trong SketchUp để export OUTPUT/IMAGES",
            item_code or "",
            item_name or "",
            text_join(envelope),
            text_join(readiness),
            str(ruby_path.relative_to(project_root)) if ruby_path else "CHƯA CÓ RUBY",
            source_text(source) or text_join(source),
        ]
        for col, value in enumerate(values, start=1):
            cell = items_ws.cell(index, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TABLE_BORDER
        items_ws.row_dimensions[index].height = 95
        if image_path:
            if not add_image(items_ws, image_path, f"A{index}"):
                items_ws.cell(index, 1, f"Không nhúng được ảnh: {image_path.name}")
                items_ws.cell(index, 1).fill = REVIEW_FILL
        else:
            items_ws.cell(index, 1).fill = REVIEW_FILL
        if not ruby_path:
            items_ws.cell(index, 6).fill = REVIEW_FILL

    mat_ws = wb.create_sheet("VAT_LIEU")
    mat_headers = [
        "Mã vật liệu", "Tên / mô tả", "Dày mm", "Diện tích net m²", "Diện tích yêu cầu m²",
        "Thể tích m³", "Khổ tấm", "Số tấm lý thuyết", "Item liên quan", "Trạng thái"
    ]
    mat_rows = []
    for row in bom_rows:
        material = material_key(row)
        mat_rows.append([
            material,
            first_value(row, "material_name", "description", "name") or "",
            row.get("thickness_mm"),
            row.get("net_area_m2"),
            row.get("required_area_m2"),
            row.get("net_volume_m3"),
            f"{row.get('sheet_length_mm') or ''} x {row.get('sheet_width_mm') or ''}".strip(" x"),
            row.get("theoretical_sheet_count"),
            text_join(row.get("item_ids")),
            first_value(row, "status", "readiness") or "THEO_DU_LIEU_HIEN_CO",
        ])
    if not mat_rows:
        mat_rows.append(["", "CHƯA CÓ BOM — workbook vẫn được tạo để tránh mất deliverable", "", "", "", "", "", "", "", "REVIEW"])
    write_table(mat_ws, mat_headers, mat_rows)
    set_widths(mat_ws, {"A": 20, "B": 40, "C": 12, "D": 16, "E": 18, "F": 16, "G": 20, "H": 18, "I": 35, "J": 24})

    detail_ws = wb.create_sheet("CHI_TIET_VAT_LIEU")
    detail_headers = ["Hạng mục", "Region / Part / Layer", "Vật liệu", "Vai trò", "Kích thước / bounds", "Trạng thái", "Nguồn"]
    detail_rows = []
    for row in material_regions:
        detail_rows.append([
            first_value(row, "item_code", "item_id", "host_item", "id") or "",
            first_value(row, "region", "part", "layer", "surface", "name") or "",
            first_value(row, "material_code", "material_name", "material") or "",
            first_value(row, "material_role", "role", "type") or "",
            text_join(first_value(row, "bounds", "dimensions", "geometry")),
            first_value(row, "status", "readiness", "derivation_state") or "",
            source_text(first_value(row, "source", "evidence")) or text_join(first_value(row, "source", "evidence")),
        ])
    if not detail_rows:
        detail_rows.append(["", "", "", "", "", "REVIEW", "CHƯA CÓ material-regions.json"])
    write_table(detail_ws, detail_headers, detail_rows)
    set_widths(detail_ws, {"A": 18, "B": 30, "C": 32, "D": 22, "E": 34, "F": 22, "G": 65})

    supplier_ws = wb.create_sheet("NHA_CUNG_CAP")
    supplier_headers = [
        "Vật liệu", "Nhà cung cấp", "Thương hiệu", "Mã sản phẩm", "Quy cách", "Đơn giá",
        "Đơn vị", "Khu vực", "Website", "Ngày kiểm tra", "Source URL", "Xác minh"
    ]
    supplier_table = []
    for row in suppliers:
        supplier_table.append([
            first_value(row, "material_code", "material", "material_name") or "",
            first_value(row, "supplier_name", "supplier", "name") or "",
            row.get("brand"), row.get("product_code"), first_value(row, "spec", "specification", "size"),
            first_value(row, "unit_price_vnd", "unit_price", "price"), row.get("unit"),
            first_value(row, "area", "region", "location"), row.get("website"),
            first_value(row, "checked_date", "checked_utc", "date"), row.get("source_url"),
            first_value(row, "verification_status", "status") or "UNVERIFIED",
        ])
    if not supplier_table:
        supplier_table.append(["", "", "", "", "", "", "", "", "", "", "", "SUPPLIER_NOT_VERIFIED"])
    write_table(supplier_ws, supplier_headers, supplier_table)
    set_widths(supplier_ws, {"A": 24, "B": 28, "C": 20, "D": 18, "E": 30, "F": 16, "G": 12, "H": 20, "I": 35, "J": 18, "K": 55, "L": 22})

    review_ws = wb.create_sheet("REVIEW")
    review_headers = ["Hạng mục", "Vấn đề", "Mức ảnh hưởng", "Nguồn", "Trạng thái"]
    review_table = []
    for row in review_rows:
        review_table.append([
            first_value(row, "item_id", "item_code", "id") or "",
            text_join(first_value(row, "reasons", "reason", "issue", "message")),
            first_value(row, "severity", "impact") or "",
            source_text(row.get("source")),
            first_value(row, "status", "readiness") or "OPEN",
        ])
    if not review_table:
        review_table.append(["", "Không có review record được xuất", "", "", "NONE"])
    write_table(review_ws, review_headers, review_table)
    set_widths(review_ws, {"A": 18, "B": 65, "C": 20, "D": 60, "E": 18})

    source_ws = wb.create_sheet("SOURCE")
    source_rows = []
    for row in items:
        src = row.get("source")
        if src:
            source_rows.append([first_value(row, "item_code", "item_id", "id") or "", source_text(src)])
    write_table(source_ws, ["Hạng mục", "Evidence / source"], source_rows or [["", "Xem WORK/manifests/input-manifest.json và reports"]])
    set_widths(source_ws, {"A": 20, "B": 100})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def supplier_price_map(suppliers: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in suppliers:
        key = material_key(row)
        price = first_value(row, "unit_price_vnd", "unit_price", "price")
        status = str(first_value(row, "verification_status", "status") or "").upper()
        if price in (None, ""):
            continue
        current = result.get(key)
        if current is None or status in {"VERIFIED", "CONFIRMED", "READY"}:
            result[key] = row
    return result


def build_quote_workbook(
    project_root: Path,
    project_meta: dict[str, Any],
    bom_rows: list[dict[str, Any]],
    suppliers: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BAO_GIA"
    ws["A1"] = "AI-dg — BÁO GIÁ SƠ BỘ / CÓ KIỂM SOÁT NGUỒN"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = project_meta.get("project_name", project_root.name)
    ws["A3"] = "Giá trống nghĩa là chưa có nguồn giá xác minh; AI-dg không tự bịa đơn giá."
    ws["A3"].fill = REVIEW_FILL

    headers = ["Mã VL", "Vật liệu / mô tả", "Quy cách", "ĐVT", "Khối lượng", "Đơn giá VND", "Thành tiền VND", "Trạng thái giá", "Nguồn giá"]
    price_map = supplier_price_map(suppliers)
    rows = []
    for row in bom_rows:
        key = material_key(row)
        supplier = price_map.get(key, {})
        qty = row.get("theoretical_sheet_count")
        unit = "tấm" if qty not in (None, "") else "m²"
        if qty in (None, ""):
            qty = first_value(row, "required_area_m2", "net_area_m2")
        price = first_value(row, "unit_price_vnd", "unit_price", "price")
        if price in (None, ""):
            price = first_value(supplier, "unit_price_vnd", "unit_price", "price")
        price_status = first_value(supplier, "verification_status", "status") if supplier else "CHUA_CO_DON_GIA"
        price_source = first_value(supplier, "source_url", "website") if supplier else ""
        spec = f"dày={row.get('thickness_mm') or '?'} mm; khổ={row.get('sheet_length_mm') or '?'}x{row.get('sheet_width_mm') or '?'}"
        rows.append([key, first_value(row, "material_name", "description") or key, spec, unit, qty, price, None, price_status, price_source])

    if not rows:
        rows.append(["", "CHƯA CÓ BOM / KHỐI LƯỢNG", "", "", "", "", None, "REVIEW", ""])

    start_row = 5
    write_table(ws, headers, rows, start_row=start_row)
    for r in range(start_row + 1, start_row + 1 + len(rows)):
        ws.cell(r, 7, f'=IF(OR(E{r}="",F{r}=""),"",E{r}*F{r})')
        ws.cell(r, 6).number_format = '#,##0'
        ws.cell(r, 7).number_format = '#,##0'
        if ws.cell(r, 8).value in ("CHUA_CO_DON_GIA", "UNVERIFIED", ""):
            ws.cell(r, 8).fill = REVIEW_FILL

    summary_start = start_row + len(rows) + 3
    ws.cell(summary_start, 5, "Tổng vật liệu").font = BOLD_FONT
    ws.cell(summary_start, 7, f"=SUM(G{start_row + 1}:G{start_row + len(rows)})")
    ws.cell(summary_start + 1, 5, "Nhân công (nhập tay nếu có)").font = BOLD_FONT
    ws.cell(summary_start + 1, 7, "")
    ws.cell(summary_start + 2, 5, "Vận chuyển / phụ phí (nhập tay)").font = BOLD_FONT
    ws.cell(summary_start + 2, 7, "")
    ws.cell(summary_start + 3, 5, "Lợi nhuận % (nhập tay)").font = BOLD_FONT
    ws.cell(summary_start + 3, 7, "")
    ws.cell(summary_start + 4, 5, "TỔNG BÁO GIÁ").font = TITLE_FONT
    ws.cell(summary_start + 4, 7, f'=IF(G{summary_start}="","",(G{summary_start}+N(G{summary_start + 1})+N(G{summary_start + 2}))*(1+N(G{summary_start + 3})))')
    for r in range(summary_start, summary_start + 5):
        ws.cell(r, 7).number_format = '#,##0'

    set_widths(ws, {"A": 20, "B": 38, "C": 32, "D": 12, "E": 15, "F": 18, "G": 20, "H": 24, "I": 55})

    supplier_ws = wb.create_sheet("DON_GIA_NGUON")
    supplier_headers = ["Vật liệu", "Nhà cung cấp", "Đơn giá", "ĐVT", "Quy cách", "Website", "Source URL", "Ngày kiểm tra", "Xác minh"]
    supplier_rows = []
    for row in suppliers:
        supplier_rows.append([
            material_key(row), first_value(row, "supplier_name", "supplier", "name") or "",
            first_value(row, "unit_price_vnd", "unit_price", "price"), row.get("unit"),
            first_value(row, "spec", "specification", "size") or "", row.get("website"),
            row.get("source_url"), first_value(row, "checked_date", "checked_utc", "date"),
            first_value(row, "verification_status", "status") or "UNVERIFIED",
        ])
    write_table(supplier_ws, supplier_headers, supplier_rows or [["", "", "", "", "", "", "", "", "SUPPLIER_NOT_VERIFIED"]])
    set_widths(supplier_ws, {"A": 24, "B": 28, "C": 16, "D": 12, "E": 30, "F": 35, "G": 55, "H": 20, "I": 22})

    review_ws = wb.create_sheet("REVIEW")
    review_table = [[
        first_value(row, "item_id", "item_code", "id") or "",
        text_join(first_value(row, "reasons", "reason", "issue", "message")),
        first_value(row, "status", "readiness") or "OPEN",
    ] for row in review_rows]
    write_table(review_ws, ["Hạng mục", "Vấn đề", "Trạng thái"], review_table or [["", "Không có review record", "NONE"]])
    set_widths(review_ws, {"A": 18, "B": 75, "C": 20})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    root = args.project_root.expanduser().resolve()
    marker = root / "project.ai-dg.json"
    if not marker.is_file():
        raise SystemExit(f"AI-dg project marker not found: {marker}")

    meta = load_json(marker, {})
    takeoff = root / "OUTPUT" / "TAKEOFF"
    geometry_root = root / "WORK" / "geometry"
    images_root = root / "OUTPUT" / "IMAGES"
    excel_root = root / "OUTPUT" / "EXCEL"

    items_payload = load_json(takeoff / "items.json", {})
    material_payload = load_json(takeoff / "material-regions.json", {})
    bom_payload = load_json(takeoff / "bom.json", {})
    suppliers_payload = load_json(takeoff / "suppliers.json", {})
    review_payload = load_json(takeoff / "review-queue.json", {})
    geometry_payload = load_json(geometry_root / "geometry-ledger.json", {})

    items = payload_rows(items_payload, ["items", "rows"])
    material_regions = payload_rows(material_payload, ["material_regions", "regions", "items"])
    bom_rows = payload_rows(bom_payload, ["bom", "materials", "rows"])
    suppliers = payload_rows(suppliers_payload, ["suppliers", "rows"])
    review_rows = payload_rows(review_payload, ["review", "queue", "items"])
    if not review_rows:
        review_rows = payload_rows(bom_payload, ["review"])
    geometry_rows = payload_rows(geometry_payload, ["items", "item_ledgers", "geometry_ledgers"])
    images = image_index(images_root)

    material_path = excel_root / "AI-dg_Tong-hop-vat-lieu.xlsx"
    quote_path = excel_root / "AI-dg_Bao-gia.xlsx"

    build_material_workbook(root, meta if isinstance(meta, dict) else {}, items, geometry_rows, material_regions, bom_rows, suppliers, review_rows, images, material_path)
    build_quote_workbook(root, meta if isinstance(meta, dict) else {}, bom_rows, suppliers, review_rows, quote_path)

    print(f"Material workbook: {material_path}")
    print(f"Quotation workbook: {quote_path}")
    print(f"Images available: {len(images)}")
    if not suppliers:
        print("Supplier status: SUPPLIER_NOT_VERIFIED (no suppliers.json)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

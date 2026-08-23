#!/usr/bin/env python3
"""Enrich AI-dg material workbook with drawing-backed material specifications.

Reads:
  OUTPUT/TAKEOFF/material-specifications.json
  OUTPUT/EXCEL/AI-dg_Tong-hop-vat-lieu.xlsx

Writes/updates:
  sheet THONG_SO_VAT_LIEU
  optional synthesized spec text in VAT_LIEU when a material key can be matched

Never invents missing values. Unknown values remain blank/UNKNOWN.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Material Excel enrichment requires openpyxl. Install AI-dg runtime extras."
    ) from exc

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Enrich AI-dg material Excel with specifications")
    parser.add_argument("project_root", type=Path)
    return parser.parse_args()


def load_json(path: Path) -> Any:
    if not path.is_file():
        raise SystemExit(f"Material specification file missing: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SystemExit(f"Invalid material specification JSON: {path}: {exc}") from exc


def rows_from_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("materials", "material_specifications", "rows"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
    return []


def text(value: Any, unknown: bool = False) -> str:
    if value in (None, "", []):
        return "UNKNOWN" if unknown else ""
    if isinstance(value, list):
        return "; ".join(text(v) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={text(v)}" for k, v in value.items())
    return str(value)


def source_text(row: dict[str, Any]) -> str:
    value = row.get("sources") or row.get("source") or row.get("evidence")
    return text(value)


def spec_text(row: dict[str, Any]) -> str:
    parts = []
    ordered = [
        ("Vật liệu", row.get("material_name") or row.get("material_family") or row.get("core")),
        ("Cốt", row.get("core")),
        ("Hoàn thiện", row.get("finish")),
        ("Màu", row.get("color")),
        ("Dày", f"{row.get('thickness_mm')} mm" if row.get("thickness_mm") not in (None, "") else None),
        ("Loại kính", row.get("glass_type")),
        ("Decal/film", row.get("film_decal") or row.get("decal") or row.get("film")),
        ("Xử lý cạnh", row.get("edge_treatment")),
        ("Keo/Sealant", row.get("adhesive_sealant") or row.get("sealant") or row.get("adhesive")),
        ("Quy cách khác", row.get("spec_text") or row.get("specification")),
    ]
    seen = set()
    for label, value in ordered:
        if value in (None, "", []):
            continue
        rendered = text(value)
        key = (label, rendered)
        if key in seen:
            continue
        seen.add(key)
        parts.append(f"{label}: {rendered}")
    return " | ".join(parts)


def slug(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def main() -> int:
    args = parse_args()
    root = args.project_root.expanduser().resolve()
    specs_path = root / "OUTPUT" / "TAKEOFF" / "material-specifications.json"
    workbook_path = root / "OUTPUT" / "EXCEL" / "AI-dg_Tong-hop-vat-lieu.xlsx"

    if not workbook_path.is_file():
        raise SystemExit(f"Material workbook missing: {workbook_path}")

    rows = rows_from_payload(load_json(specs_path))
    wb = load_workbook(workbook_path)

    if "THONG_SO_VAT_LIEU" in wb.sheetnames:
        del wb["THONG_SO_VAT_LIEU"]
    ws = wb.create_sheet("THONG_SO_VAT_LIEU", 3 if len(wb.sheetnames) >= 3 else len(wb.sheetnames))

    headers = [
        "Hạng mục", "Material ID", "Vật liệu / hệ vật liệu", "Vai trò", "Region / Part / Layer",
        "Cốt / Core", "Hoàn thiện", "Màu", "Dày mm", "Loại kính", "Decal / Film",
        "Xử lý cạnh", "Keo / Sealant", "Thông số tổng hợp", "Trạng thái", "Nguồn"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "CHƯA CÓ MATERIAL SPECIFICATION RECORD", "REVIEW", str(specs_path)])
        for cell in ws[2]:
            cell.fill = REVIEW_FILL
    else:
        for row in rows:
            ws.append([
                text(row.get("item_code") or row.get("item_id") or row.get("host_item")),
                text(row.get("material_id") or row.get("material_code") or row.get("id")),
                text(row.get("material_name") or row.get("material_family") or row.get("material")),
                text(row.get("role") or row.get("material_role")),
                text(row.get("host_region") or row.get("region") or row.get("part") or row.get("layer") or row.get("surface")),
                text(row.get("core"), unknown=True),
                text(row.get("finish"), unknown=True),
                text(row.get("color"), unknown=True),
                text(row.get("thickness_mm"), unknown=True),
                text(row.get("glass_type"), unknown=True),
                text(row.get("film_decal") or row.get("decal") or row.get("film"), unknown=True),
                text(row.get("edge_treatment"), unknown=True),
                text(row.get("adhesive_sealant") or row.get("sealant") or row.get("adhesive"), unknown=True),
                spec_text(row),
                text(row.get("status") or row.get("readiness") or row.get("derivation_state")),
                source_text(row),
            ])

    widths = [16, 22, 34, 24, 30, 18, 20, 18, 12, 24, 28, 28, 24, 70, 20, 75]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Best-effort enrichment of existing VAT_LIEU sheet without changing its existing data.
    if "VAT_LIEU" in wb.sheetnames:
        mat_ws = wb["VAT_LIEU"]
        new_col = mat_ws.max_column + 1
        mat_ws.cell(1, new_col, "Thông số vật liệu từ bản vẽ")
        mat_ws.cell(1, new_col).fill = HEADER_FILL
        mat_ws.cell(1, new_col).font = HEADER_FONT
        mat_ws.cell(1, new_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        mat_ws.column_dimensions[mat_ws.cell(1, new_col).column_letter].width = 70

        spec_index: dict[str, list[str]] = {}
        for row in rows:
            keys = {
                slug(row.get("material_id")), slug(row.get("material_code")), slug(row.get("material_name")),
                slug(row.get("material_family")), slug(row.get("core")),
            }
            rendered = spec_text(row)
            for key in {k for k in keys if k}:
                spec_index.setdefault(key, []).append(rendered)

        for r in range(2, mat_ws.max_row + 1):
            candidates = [slug(mat_ws.cell(r, 1).value), slug(mat_ws.cell(r, 2).value)]
            matches: list[str] = []
            for candidate in candidates:
                if not candidate:
                    continue
                for key, values in spec_index.items():
                    if candidate == key or candidate in key or key in candidate:
                        matches.extend(values)
            unique = []
            for value in matches:
                if value and value not in unique:
                    unique.append(value)
            if unique:
                mat_ws.cell(r, new_col, "\n".join(unique))
            else:
                mat_ws.cell(r, new_col, "KHÔNG KHỚP ĐƯỢC SPEC — xem THONG_SO_VAT_LIEU")
                mat_ws.cell(r, new_col).fill = REVIEW_FILL
            mat_ws.cell(r, new_col).alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(workbook_path)
    print(f"Material workbook enriched -> {workbook_path}")
    print(f"Material specification rows: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

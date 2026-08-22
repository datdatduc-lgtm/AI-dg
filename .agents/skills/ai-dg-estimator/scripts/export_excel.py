#!/usr/bin/env python3
"""Export validated AI-dg items/BOM JSON into a review-friendly Excel workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export AI-dg data to Excel")
    parser.add_argument("items", type=Path)
    parser.add_argument("bom", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 40)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 10)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def source_text(source: dict[str, Any]) -> str:
    return f"{source.get('pdf', '')} | page {source.get('page', '')} | {source.get('evidence', '')}"


def main() -> int:
    args = parse_args()
    items_payload = json.loads(args.items.read_text(encoding="utf-8"))
    bom_payload = json.loads(args.bom.read_text(encoding="utf-8"))
    items = items_payload.get("items", [])
    bom = bom_payload.get("bom", [])
    review = bom_payload.get("review", [])

    wb = Workbook()
    summary = wb.active
    summary.title = "SUMMARY"
    summary["A1"] = "AI-dg Estimation Summary"
    summary["A1"].font = TITLE_FONT
    summary.append([])
    summary.append(["Metric", "Value"])
    summary.append(["Project", items_payload.get("project", "")])
    summary.append(["Extracted item records", len(items)])
    summary.append(["BOM groups", len(bom)])
    summary.append(["Review records", len(review)])
    summary.append(["Scope", "V0.1 takeoff only; not a final quotation"])
    for cell in summary[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 55

    items_ws = wb.create_sheet("ITEMS")
    item_headers = [
        "ID", "Item Code", "Item Name", "Part", "Source Material", "Material",
        "Length mm", "Width mm", "Thickness mm", "Qty", "Confidence",
        "Review Required", "Source PDF", "Page", "Evidence"
    ]
    item_rows = []
    for item in items:
        src = item.get("source", {})
        item_rows.append([
            item.get("id"), item.get("item_code"), item.get("item_name"), item.get("part_name"),
            item.get("source_material_code"), item.get("material_code"), item.get("length_mm"),
            item.get("width_mm"), item.get("thickness_mm"), item.get("quantity"),
            item.get("confidence"), item.get("review_required"), src.get("pdf"), src.get("page"),
            src.get("evidence")
        ])
    write_table(items_ws, item_headers, item_rows)
    items_ws.column_dimensions["O"].width = 40

    bom_ws = wb.create_sheet("BOM")
    bom_headers = [
        "Material", "Thickness mm", "Net Area m2", "Waste Factor", "Required Area m2",
        "Net Volume m3", "Sheet L mm", "Sheet W mm", "Theoretical Sheets", "Item IDs"
    ]
    bom_rows = []
    for row in bom:
        bom_rows.append([
            row.get("material_code"), row.get("thickness_mm"), row.get("net_area_m2"),
            row.get("waste_factor"), row.get("required_area_m2"), row.get("net_volume_m3"),
            row.get("sheet_length_mm"), row.get("sheet_width_mm"), row.get("theoretical_sheet_count"),
            ", ".join(row.get("item_ids", []))
        ])
    write_table(bom_ws, bom_headers, bom_rows)

    review_ws = wb.create_sheet("REVIEW")
    review_headers = ["Item ID", "Reasons", "Source"]
    review_rows = [
        [row.get("item_id"), "; ".join(row.get("reasons", [])), source_text(row.get("source", {}))]
        for row in review
    ]
    write_table(review_ws, review_headers, review_rows)
    review_ws.column_dimensions["B"].width = 45
    review_ws.column_dimensions["C"].width = 60

    sources_ws = wb.create_sheet("SOURCES")
    source_rows = []
    seen: set[tuple] = set()
    for item in items:
        src = item.get("source", {})
        key = (src.get("pdf"), src.get("page"), src.get("evidence"))
        if key in seen:
            continue
        seen.add(key)
        source_rows.append([src.get("pdf"), src.get("page"), src.get("evidence"), item.get("id")])
    write_table(sources_ws, ["PDF", "Page", "Evidence", "First Item ID"], source_rows)
    sources_ws.column_dimensions["C"].width = 60

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Excel workbook written -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
